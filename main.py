from qtmodules import *

from vues import RouteView, CountryView, WorldView
import sys
from controller import Controller
from graphs import Grapher
import logging
from utils import *
from openpyxl import Workbook
from PyQt6.QtWidgets import QFileDialog

# Import to update the country list only once on all instances
from widgets import CountrySelect, SplashScreen

logging.basicConfig(level=logging.INFO)


class MainWindow(QMainWindow):
    def __init__(self, password, delay=500):
        super().__init__()

        self.c = Controller(password)

        self.setWindowTitle("Système d'onglets")
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        # oui, on a rajouté des attentes arbitraitres pour que le splash screen soit visible
        # parce que après tout, il est quand même joli
        # ça serait dommage de le cacher trop vite ;)

        self.splash = SplashScreen("Construction de l'interface...")
        self.splash.show()
        self.create_tabs()
        self.splash.block_thread_for(delay)

        self.splash.showMessage("Chargement des onglets...")
        self.connect_tabs()
        self.splash.block_thread_for(delay)

        self.splash.showMessage("Initialisation des graphiques...")
        self.grapher = Grapher()
        self.splash.block_thread_for(delay * 3/5)

        self.splash.showMessage("Remplissage depuis \nla base de données...")
        self.populate_tabs()
        self.splash.block_thread_for(delay)

        self.splash.showMessage(
            "Bienvue sur notre outil de \nvisualisation de données!", animated=True)

        self.splash.destroy_in(delay*4)
        self.splash.timer.timeout.connect(self.show)
        self.splash.timer.timeout.connect(self.raise_)

    def create_tabs(self):
        """Create the tabs and add them to the tab widget"""

        # TAB 1
        self.world_widget = WorldView()
        self.tab_widget.addTab(self.world_widget, "Aperçu Aeroports")
        self.tab_widget.setStyleSheet("QTabWidget::pane {border: 0;}")
        self.tab_widget.setTabIcon(0, QIcon("img/Monde.png"))

        # TAB 2
        self.country_view = CountryView()
        self.tab_widget.addTab(self.country_view, "Aperçu Pays")
        self.tab_widget.setStyleSheet("QTabWidget::pane {border: 0;}")
        self.tab_widget.setTabIcon(1, QIcon("img/Pays.png"))

        # TAB 3
        self.route_view = RouteView()
        self.tab_widget.addTab(self.route_view, "Compagnies Vertueuses")
        self.tab_widget.setStyleSheet("QTabWidget::pane {border: 0;}")
        self.tab_widget.setTabIcon(2, QIcon("img/comparaison.png"))

    def connect_tabs(self):
        """Connect the tabs to the controller and to each other"""
        # TAB 1
        self.world_widget.country_selected.connect(
            self.world_view_show_airports)
        self.world_widget.airport_selected.connect(
            self.world_view_update_graph)

        # TAB 2
        self.country_view.country_selected.connect(
            self.country_view_update_graphs)

        # TAB 3
        self.route_view.country_choiced_left.connect(
            self.route_view_update_airport_list_left)
        self.route_view.country_choiced_right.connect(
            self.route_view_update_airport_list_right)
        self.route_view.aero_choiced.connect(self.route_view_update_data)
        self.route_view.clicked_dowload.connect(self.download_button_clicked)

    def populate_tabs(self):
        """Populate the tabs with data from the database"""
        self.populate_country_list()

    # CONNECTIONS / SIGNALS

    # TAB 1

    def world_view_show_airports(self, country: str):
        # gather airport data
        data = self.c.get_useful_airports_coords_by_country(country)
        logging.info(f"{len(data)} airports loaded for {country}")
        print(data)
        self.world_widget.map_widget.load_marker_list(data)

        # also update the graph
        self.world_widget.graph_widget.load_text(
            "Veuillez choisir un aéroport sur la carte pour afficher les routes les plus polluantes partant de celui-ci",
        )

    def world_view_update_graph(self, lon, lat):

        # get the airport id
        airport_id = self.c.get_airport_id_by_coords(lon, lat)
        airport_name = self.c.get_airport_name_by_id(airport_id)
        tuples = self.c.get_most_polluting_routes(airport_id, 5)

        if not tuples:
            logging.error("No data to display")
            self.world_widget.graph_widget.load_text(
                f"Pas de données à afficher pour l'aéroport {airport_name}")
            return

        data, labels = zip(*tuples)

        # TODO - The graph can sometimes show twice the same airport

        # generate the graph
        file = self.grapher.pie_chart_graph(
            data, labels, f"Routes les plus polluantes \npartant de {airport_name}", "Pollution (kg CO2)")

        # display the graph
        self.world_widget.graph_widget.load_image(file)

    # TAB 2

    def country_view_update_graphs(self, country: str):

        # GRAPH 1 : Part vol intérieur / vol international

        # Get airport info
        inside_flights, outside_flights = self.c.get_country_flights_info(
            country)

        # Update graph
        file = self.grapher.pie_chart_graph(
            [inside_flights, outside_flights],
            ["Vols intérieurs", "Vols internationaux"],
            f"Répartition des vols \nde {country}",
            show_percent=True,
        )
        self.country_view.graph_widget1.load_image(file)

        # GRAPH 2 : Top 5 des pays destination les plus courants

        # Get country info
        data = self.c.get_top_countries(country)
        country_flights, country_name = zip(*data)
        total_flights = self.c.get_total_flights(country)
        country_flights = [round(x/total_flights*100, 2)
                           for x in country_flights]

        # Update graph
        file = self.grapher.funnel_chart_graph(
            country_flights, country_name, f"Pourcentage des vols vers\n l'étranger depuis {country}", data_is_percent=True
        )
        self.country_view.graph_widget2.load_image(file)

    # TAB 3

    def route_view_update_airport_list_left(self, country):
        if self.route_view.right_airport_selected == None:
            self.route_view.update_leftcombobox(
                self.c.get_aeroports_by_country(country))
        else:
            liste_aero = self.c.get_aeroports_by_country_constraint(
                country, self.route_view.right_airport_selected)
            self.route_view.update_leftcombobox(liste_aero)
            if liste_aero == []:
                self.route_view.best_route_widget.reset()

    def route_view_update_airport_list_right(self, country):
        if self.route_view.left_airport_selected == None:
            self.route_view.update_rightcombobox(
                self.c.get_aeroports_by_country(country))
        else:
            liste_aero = self.c.get_aeroports_by_country_constraint(
                country, self.route_view.left_airport_selected)
            self.route_view.update_rightcombobox(liste_aero)
            if liste_aero == []:
                self.route_view.best_route_widget.reset()
                self.route_view.map_widget.reset_map()

    def route_view_update_data(self, airport1, airport2):
        """Update the data in route_view. airport1 and airport2 are the names of the airports"""

        # top widget
        distance = self.c.get_distance_between_airports(airport1, airport2)

        data = self.c.get_best_airplane_model_for_route(
            airport1, airport2)

        if data:
            airplane_model, co2_emission = data
        else:
            airplane_model = "Aucun"
            co2_emission = "Aucun"

        airlines = self.c.get_companies_for_route_by_plane_model(
            airport1, airport2, airplane_model)

        if not airlines:
            airlines = ["Aucune"]

        print(airlines)

        self.route_view.best_route_widget.set_airplane_model(airplane_model)
        self.route_view.best_route_widget.set_co2(co2_emission)
        self.route_view.best_route_widget.set_companies(airlines)
        self.route_view.best_route_widget.set_distance(distance)

        self.route_view.best_route_widget.update()

        # don't update the map if there is no data
        if not data:
            # Show an empty map instead of keeping the previous one
            self.route_view.map_widget.reset_map()
            return

        # map widget
        # get the coordinates of the airports
        coord1 = self.c.get_airport_coords_by_name(airport1)
        coord2 = self.c.get_airport_coords_by_name(airport2)
        # show the markers
        self.route_view.map_widget.load_marker_list(
            [(airport1, coord1), (airport2, coord2)], drop_current_map=True)
        # show the route
        self.route_view.map_widget.trace_path(
            coord1, coord2, drop_current_map=False)

    # POPULATE TABS
    def populate_country_list(self):
        data = self.c.get_countries_with_iso()
        CountrySelect.set_country_list(data)
        logging.debug("Country list populated")

    def download_button_clicked(self, country1, country2):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le fichier Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            self.download(country1, country2, file_path)

    def download(self, country1, country2, file_path):
        headers = ['Airline_Name', 'Departure_Airport',
                   'Destination_Airport', 'Plane', 'Seat', 'Distance', 'CO2_Emissions']
        values = self.c.get_info_download(country1, country2)

        workbook = Workbook()
        sheet = workbook.active

        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row, row_data in enumerate(values, start=2):
            for col, value in enumerate(row_data, start=1):
                sheet.cell(row=row, column=col, value=value)

        workbook.save(file_path)


if __name__ == "__main__":
    try:
        with open('Data.json', 'r') as Data:
            data = json.load(Data)
            mdp_base = hex_to_string(data['BPASS'])
    except FileNotFoundError:
        with open('Data.json', 'w') as Data:
            pass
        mdp_base = Init()
    except json.decoder.JSONDecodeError:
        mdp_base = Init()

    logo_path = "img/logo_Ccorp.png"

    import platform
    if platform.system() == 'Windows':
        import ctypes
        myappid = 'calaiscorp.calaiscologie.version1-0'  # arbitrary string
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    elif platform.system() == 'Linux':
        # TODO - implement linux taskbar icon
        logging.warning("Linux taskbar icon not implemented yet")
    elif platform.system() == 'Darwin':
        # TODO - implement mac taskbar icon
        logging.warning("Mac taskbar icon not implemented yet")

    app = QApplication(sys.argv)

    window = MainWindow(mdp_base, delay=800)
    width: int = app.primaryScreen().size().width()
    height: int = app.primaryScreen().size().height()
    window.setFixedSize(int(width/1.2), int(height/1.2))
    window.setWindowTitle("CalaisCologie")
    window.setWindowIcon(QIcon(logo_path))
    window.setStyleSheet(
        "MainWindow{ background-image: url(img/aurore.png); background-repeat: no-repeat; background-position: center; }")

    app.exec()
